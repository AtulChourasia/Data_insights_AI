import pandas as pd
import logging
from typing import List, Dict, Any, Optional, Tuple
import io
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def get_excel_sheets(file_path: str) -> List[Dict[str, Any]]:
    """
    Get information about all sheets in an Excel file.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        List of dictionaries containing sheet information
    """
    try:
        # Use openpyxl for better performance with large files
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Get dimensions (approximate for large files)
            dimensions = ws.calculate_dimension()
            
            sheets.append({
                "name": sheet_name,
                "dimensions": dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "active": wb.active.title == sheet_name
            })
            
        return sheets
        
    except Exception as e:
        logger.error(f"Error reading Excel file {file_path}: {e}")
        raise


def read_excel_sheet(file_path: str, sheet_name: str, nrows: int = 5) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Read a specific sheet from an Excel file.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read
        nrows: Number of rows to read (for preview)
        
    Returns:
        Tuple of (dataframe, metadata)
    """
    try:
        # Read the Excel file
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            nrows=nrows,
            engine='openpyxl'
        )
        
        # Get metadata
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[sheet_name]
        
        metadata = {
            "sheet_name": sheet_name,
            "total_rows": ws.max_row,
            "total_columns": ws.max_column,
            "headers": list(df.columns) if not df.empty else [],
            "sample_data": df.head().to_dict(orient='records')
        }
        
        return df, metadata
        
    except Exception as e:
        logger.error(f"Error reading sheet {sheet_name} from {file_path}: {e}")
        raise


def detect_headers(df: pd.DataFrame, n_rows_to_check: int = 5) -> Tuple[int, List[str]]:
    """
    Detect the header row in a DataFrame.
    
    Args:
        df: Input DataFrame
        n_rows_to_check: Number of rows to check for header detection
        
    Returns:
        Tuple of (header_row_index, header_columns)
    """
    # Check if first row looks like a header
    first_row = df.iloc[0].values
    if all(isinstance(x, str) for x in first_row):
        return 0, list(first_row)
    
    # If not, try to find a row that looks like a header
    for i in range(1, min(n_rows_to_check, len(df))):
        row = df.iloc[i].values
        if all(isinstance(x, str) for x in row):
            return i, list(row)
    
    # If no header found, use default column names
    return 0, [f"Column_{i+1}" for i in range(len(df.columns))]


def clean_column_names(columns: List[str]) -> List[str]:
    """
    Clean and standardize column names.
    
    Args:
        columns: List of column names
        
    Returns:
        List of cleaned column names
    """
    cleaned = []
    for i, col in enumerate(columns):
        if not col or pd.isna(col):
            cleaned.append(f"column_{i+1}")
            continue
            
        # Convert to string and clean
        col = str(col).strip().lower()
        col = col.replace(' ', '_').replace('-', '_').replace('.', '_')
        col = ''.join(c if c.isalnum() or c == '_' else '' for c in col)
        col = col.strip('_')
        
        if not col:
            col = f"column_{i+1}"
            
        # Ensure unique column names
        if col in cleaned:
            j = 1
            while f"{col}_{j}" in cleaned:
                j += 1
            col = f"{col}_{j}"
            
        cleaned.append(col)
        
    return cleaned


def excel_to_dataframes(file_path: str, header_row: int = 0) -> Dict[str, pd.DataFrame]:
    """
    Read all sheets from an Excel file into DataFrames.
    
    Args:
        file_path: Path to the Excel file
        header_row: Row index to use as header (0-based)
        
    Returns:
        Dictionary mapping sheet names to DataFrames
    """
    try:
        # Read all sheets
        xls = pd.ExcelFile(file_path, engine='openpyxl')
        sheets = {}
        
        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(
                    xls,
                    sheet_name=sheet_name,
                    header=header_row,
                    engine='openpyxl'
                )
                
                # Clean column names
                df.columns = clean_column_names(df.columns.tolist())
                sheets[sheet_name] = df
                
            except Exception as e:
                logger.warning(f"Error reading sheet {sheet_name}: {e}")
                continue
                
        return sheets
        
    except Exception as e:
        logger.error(f"Error reading Excel file {file_path}: {e}")
        raise
